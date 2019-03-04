from openpyxl import *
import pandas as pd
from numpy import where


text_container = ""

class key_fields(object):

	def __init__(self,sap_fields):
		self.sap_fields = []

	def flag():
		pass


def trim_tables(table1, table2):

	first = table1.drop_duplicates(keep=False)
	second	= table2.drop_duplicates(keep=False)

	drop_these = get_col_diff(first, second)

	print_this("\n Removing the below columns: %s"  %drop_these)
	first = first.drop(drop_these, axis =1)

	
	return first


#count entries
def count_missing(Extract,DCW):

	DCW_rows = DCW.shape[0]
	Extract_rows = Extract.shape[0]

	print_this("\nFound %r unique entries in DCW and %r unique entries in Extract" %(DCW_rows,Extract_rows))

	if DCW_rows > Extract_rows:
		print_this("\nExpected to find at least %r missing entries " %(DCW_rows - Extract_rows))
	else:
		pass


def get_col_diff(table1,table2):
	table1_columns = list(table1)
	table2_columns = list(table2)

	set_table2 = set(table2_columns)
	col_difference = [x for x in table1_columns if x not in set_table2]

	return col_difference

def stage_tables(file,sheet_number,key):

	wb = pd.ExcelFile(file, on_demand = True)
	wb_sheetnames = wb.sheet_names

	

	table = wb.parse(sheet_name = wb_sheetnames[sheet_number - 1], header = None, skiprows = None, dtype = object)
	header_line = find_row_containing(table, key[0]) 


	table = wb.parse(sheet_name = wb_sheetnames[sheet_number - 1], header = 0, skiprows = header_line)

	#clean duplicates in Extract(table1) and DCW(table2) 
	table = table.drop_duplicates(keep=False)
	table = table.dropna(axis = 1, how = 'all')


	wb.close()

	return table


def find_row_containing(df1, value):
	return df1[df1 == value].dropna(how = 'all').index.values[0]


def print_this(string):
	
	global text_container
	text_container = text_container +  string + "\n"
	
	
#####################################################################


def main(file_names, unique_key):

	#unique field in DCW
	#unique_key = "MATNR"  as default
	#table1 being Extract
	table1 = stage_tables(file_names[0], 1, unique_key)
	#table2 being DCW
	table2 = stage_tables(file_names[1], 1, unique_key)


	table1 = trim_tables(table1,table2)
	table2 = trim_tables(table2,table1)


	#print "\n\nChecking %r entries in DCW ... " %(table2.shape[0])

	count_missing(table1,table2)

	#print table2
	#print table1

	#combine tables
	combined_table = pd.concat([table2,table1], keys = ['DCW','Extract'])

	#drop_duplicates of combined table
	nomatch = combined_table.drop_duplicates(keep=False)
	#set nomatch column order to be as DCW column order
	nomatch = nomatch[list(table2)]


	#output discrepancy to excel
	output_here = 'output.xlsx'
	output = pd.ExcelWriter(output_here)

	if nomatch.empty:
		print_this("\nNo discrepancies found. All entries in DCW found in Extract File.")
		
		quit()

	else:
		nomatch.loc['DCW'].to_excel(output, sheet_name = 'DCW Discrepancies', engine = 'xlsxwriter')
		nomatch.loc['Extract'].to_excel(output, sheet_name = 'Extract Discrepancies', engine = 'xlsxwriter')


		print_this("\nDiscrepancies between DCW and Extract file is detailed in: %r" %(output_here))

	#extract the two multi-index 
	DCW = nomatch.loc['DCW']
	Extract = nomatch.loc['Extract']


	#merge two tables into one table
	df3 = DCW.merge(Extract, how = 'outer', on = unique_key, suffixes = ['', '_Extract'], indicator = True)

	#drop irrelevant Extract entries in Merge Table
	df3 = df3.drop(df3[df3._merge == 'right_only'].index)
	#change values in _merge
	df3['_merge'] = df3['_merge'].map({'left_only': 'Missing', 'both':'Discrepancies found'})
	#rename merge
	df3.rename(columns = {'_merge':'Error Message'}, inplace= True)
	#rearrange column _merge to the front
	df3 = df3[list(df3)[-1:] + list(df3)[:-1]]



	#iterate through column pairs and highlight differences
	remove_columns = []
	col_order = []
	col_order_num = []

	for i in list(set(DCW)-set(unique_key)):
		
		df3['Mismatch_' + i] = where(df3.loc[:, i] == df3.loc[:, i + '_Extract'], ' -- ',  df3.loc[:, i + '_Extract'] )
				
		remove_columns.append(i + '_Extract')

		if (df3['Mismatch_' + i] == ' -- ' ).all(axis = 0 ,skipna = False):
				remove_columns.append('Mismatch_' + i)
				remove_columns.append(i)
		else:
			pass		


	#drop correct and irrelevant columns
	Output_trimmed = df3.drop(remove_columns, axis = 1)



	#rearrange columns so that field is next to Mismatch_field
	for i in range(0, len(list(Output_trimmed)),1):


		try:
			mismatch_loc = list(Output_trimmed).index( 'Mismatch_' + list(Output_trimmed)[i])
			col_order.append(list(Output_trimmed)[i])
			col_order.append(list(Output_trimmed)[mismatch_loc])
			col_order_num.append(i)
			col_order_num.append(mismatch_loc)

		except:
			if i in col_order_num :
				pass	
			else:
				col_order.append(list(Output_trimmed)[i])
				col_order_num.append(i)



	Output_trimmed = Output_trimmed[col_order]

	#output to Excel Sheets
	df3.to_excel(output, sheet_name = 'Merged', engine = 'xlsxwriter')
	Output_trimmed.to_excel(output, sheet_name = 'Mismatch', engine = 'xlsxwriter')


	output.save()


